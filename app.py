import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
import numpy as np
import calendar
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Function to calculate working days in a month excluding weekends
def get_working_days(year, month):
    month_start = datetime(year, month, 1)
    month_end = datetime(year, month, calendar.monthrange(year, month)[1])
    day_count = np.busday_count(month_start.date(), (month_end + timedelta(days=1)).date())
    return day_count

# Function to calculate the total work hours in the month
def calculate_total_work_hours(year, month, hours_per_day=8):
    working_days = get_working_days(year, month)
    return working_days * hours_per_day

# Function to check if a date is a weekday or weekend
def is_weekend(date):
    return date.weekday() >= 5  # 5 for Saturday, 6 for Sunday

# Function to handle Block 1 logic
def block_1(clockify_data):
    # Extract year and month from the Clockify data (assuming it's consistent for the entire dataset)
    start_date = clockify_data['Start Date'].min()
    year = start_date.year
    month = start_date.month

    # Calculate total work hours dynamically for the given month
    TOTAL_WORK_HOURS = calculate_total_work_hours(year, month)

    # Extract unique project names from the Clockify data
    unique_projects = clockify_data['Project'].unique()

    # Dynamically create the project_columns structure
    project_columns = [(f'{project} Time allocation', f'{project} Amount') for project in unique_projects]

    # Define the columns for the output template
    template_columns = ['Department', 'Manager', 'Name', 'Amount'] + [item for sublist in project_columns for item in sublist] + ['Total Percentage']

    # Group by User and Project to calculate the total hours worked on each project
    employee_project_hours = clockify_data.groupby(['User', 'Project'])['Duration (decimal)'].sum().reset_index()

    # Calculate the percentage of time allocation for each project
    employee_project_hours['Percentage'] = (employee_project_hours['Duration (decimal)'] / TOTAL_WORK_HOURS) * 100
    employee_project_hours['Percentage'] = employee_project_hours['Percentage'].round(2)  # Round to 2 decimal places

    # Create a list to hold the data for the Template sheet
    template_data_list = []
    overtime_data_list = []
    less_than_8_data_list = []

    # Populate the Template DataFrame based on the structure provided
    for user in employee_project_hours['User'].unique():
        user_group = clockify_data.loc[clockify_data['User'] == user, 'Group'].values[0]

        total_hours_worked = employee_project_hours.loc[employee_project_hours['User'] == user, 'Duration (decimal)'].sum()

        # Track overtime distribution between weekdays and weekends
        weekdays_ot = 0
        weekend_ot = 0

        # Track daily work hours to identify less than 8 hours cases
        daily_hours = clockify_data[clockify_data['User'] == user]

        for _, day_row in daily_hours.iterrows():
            if day_row['Start Date'] == day_row['End Date']:  # Only consider if Start and End Date are the same
                is_weekend_day = is_weekend(day_row['Start Date'])
                work_duration = day_row['Duration (decimal)']

                if is_weekend_day:
                    weekend_ot += work_duration  # All hours on weekend are considered overtime
                else:
                    if work_duration > 9:
                        weekdays_ot += (work_duration - 9)
                    if work_duration < 8:
                        less_than_8_data_list.append({
                            'Name': user,
                            'Department': user_group,
                            'Manager': 'Manager',
                            'Date': day_row['Start Date'].strftime('%Y-%m-%d'),
                            'Hours': round(work_duration, 2),
                        })

        # Only add to the overtime sheet if there is any overtime
        if weekdays_ot > 0 or weekend_ot > 0:
            overtime_data_list.append({
                'User': user,
                'Project': 'All Projects',
                'Weekdays OT': round(weekdays_ot, 2),
                'Weekend OT': round(weekend_ot, 2),
            })

        row = {'Department': user_group, 'Manager': 'Manager', 'Name': user, 'Amount': ''}  # Use Group as Department
        for project_time, project_amount in project_columns:
            project_name = project_time.split(' Time allocation')[0]
            project_hours = employee_project_hours.loc[(employee_project_hours['User'] == user) & (employee_project_hours['Project'] == project_name), 'Duration (decimal)'].sum()
            capped_hours = (project_hours / total_hours_worked) * TOTAL_WORK_HOURS if total_hours_worked > TOTAL_WORK_HOURS else project_hours

            # Calculate the percentage time allocation
            percentage_time_allocation = min(100, (capped_hours / TOTAL_WORK_HOURS) * 100)

            # Assign the correctly formatted percentage to the row
            row[project_time] = f"{percentage_time_allocation:.2f}%"

            # Always set the formula for the project amount if there is any time allocation
            project_time_col_letter = get_column_letter(template_columns.index(project_time) + 1)  # Get column letter
            row[project_amount] = f"={project_time_col_letter}{len(template_data_list) + 2}*D{len(template_data_list) + 2}"

        # Cap the total percentage at 100%
        row['Total Percentage'] = f"{min(100, (total_hours_worked / TOTAL_WORK_HOURS) * 100):.2f}%"
        template_data_list.append(row)

    # Convert the list to a DataFrame with the template's column structure
    output_template = pd.DataFrame(template_data_list, columns=template_columns)

    # Prepare the second sheet "Pivot Data"
    pivot_data = employee_project_hours.pivot(index='User', columns='Project', values='Duration (decimal)').fillna(0)

    # Add Total Hours and Total Days column in Pivot Data
    pivot_data['Total Hours'] = pivot_data.sum(axis=1)
    pivot_data['Total Days'] = round(pivot_data['Total Hours'] / 8)  # Assuming an 8-hour workday

    # Convert the overtime data list to a DataFrame (without 'Overtime Hours' column)
    overtime_df = pd.DataFrame(overtime_data_list)

    # Convert the less than 8 hours data list to a DataFrame
    less_than_8_df = pd.DataFrame(less_than_8_data_list)

    # Identify discrepancies where Start Date and End Date do not match
    discrepancy_df = clockify_data[clockify_data['Start Date'] != clockify_data['End Date']]

    # Remove unnecessary columns from the discrepancy data
    discrepancy_df = discrepancy_df.drop(columns=['Client', 'Description', 'Task', 'Tags', 'Start Time', 'End Time', 'Duration (decimal)'])

    # Return necessary dataframes for further processing in Block 2
    return output_template, pivot_data, overtime_df, less_than_8_df, discrepancy_df, clockify_data

# Function to handle Block 2 logic
def block_2(output_template, pivot_data, overtime_df, less_than_8_df, discrepancy_df):
    start_date = clockify_data['Start Date'].min()
    year = start_date.year
    month = start_date.month

    # Calculate total work hours dynamically for the given month
    TOTAL_WORK_HOURS = calculate_total_work_hours(year, month)

    # Extract unique project names from the Clockify data
    unique_projects = clockify_data['Project'].unique()

    # Dynamically create the project_columns structure
    project_columns = [(f'{project} Time allocation', f'{project} Amount') for project in unique_projects]

    # Define the columns for the output template
    template_columns = ['Department', 'Manager', 'Name', 'Amount'] + [item for sublist in project_columns for item in sublist] + ['Total Percentage']

    # Group by User and Project to calculate the total hours worked on each project
    employee_project_hours = clockify_data.groupby(['User', 'Project'])['Duration (decimal)'].sum().reset_index()

    # Calculate the percentage of time allocation for each project
    employee_project_hours['Percentage'] = (employee_project_hours['Duration (decimal)'] / TOTAL_WORK_HOURS) * 100
    employee_project_hours['Percentage'] = employee_project_hours['Percentage'].round(2)  # Round to 2 decimal places

    # Create a list to hold the data for the Template sheet
    # Create a list to hold the data for the Template sheet
    template_data_list = []
    overtime_data_list = []
    less_than_8_data_list = []

    # Populate the Template DataFrame based on the structure provided
    for user in employee_project_hours['User'].unique():
        user_group = clockify_data.loc[clockify_data['User'] == user, 'Group'].values[0]

        total_hours_worked = employee_project_hours.loc[employee_project_hours['User'] == user, 'Duration (decimal)'].sum()

        # Track overtime distribution between weekdays and weekends
        weekdays_ot = 0
        weekend_ot = 0

        # Track daily work hours to identify less than 8 hours cases
        daily_hours = clockify_data[clockify_data['User'] == user]

        # Group the daily_hours by date and sum the hours for each date
        daily_hours_grouped = daily_hours.groupby('Start Date')['Duration (decimal)'].sum().reset_index()

        for _, day_row in daily_hours_grouped.iterrows():
            is_weekend_day = is_weekend(day_row['Start Date'])
            work_duration = day_row['Duration (decimal)']

            if is_weekend_day:
                weekend_ot += work_duration  # All hours on weekend are considered overtime
            else:
                if work_duration > 9:
                    weekdays_ot += (work_duration - 9)
                if work_duration < 8:
                    less_than_8_data_list.append({
                        'Name': user,
                        'Department': user_group,
                        'Manager': 'Manager',
                        'Date': day_row['Start Date'].strftime('%Y-%m-%d'),
                        'Hours': round(work_duration, 2),
                    })

        # Only add to the overtime sheet if there is any overtime
        if weekdays_ot > 0 or weekend_ot > 0:
            overtime_data_list.append({
                'User': user,
                'Project': 'All Projects',
                'Weekdays OT': round(weekdays_ot, 2),
                'Weekend OT': round(weekend_ot, 2),
            })

        row = {'Department': user_group, 'Manager': 'Manager', 'Name': user, 'Amount': ''}  # Use Group as Department
        for project_time, project_amount in project_columns:
            project_name = project_time.split(' Time allocation')[0]
            project_hours = employee_project_hours.loc[(employee_project_hours['User'] == user) & (employee_project_hours['Project'] == project_name), 'Duration (decimal)'].sum()
            capped_hours = (project_hours / total_hours_worked) * TOTAL_WORK_HOURS if total_hours_worked > TOTAL_WORK_HOURS else project_hours

            # Calculate the percentage time allocation
            percentage_time_allocation = min(100, (capped_hours / TOTAL_WORK_HOURS) * 100)

            # Assign the correctly formatted percentage to the row
            row[project_time] = f"{percentage_time_allocation:.2f}%"

            # Always set the formula for the project amount if there is any time allocation
            project_time_col_letter = get_column_letter(template_columns.index(project_time) + 1)  # Get column letter
            row[project_amount] = f"={project_time_col_letter}{len(template_data_list) + 2}*D{len(template_data_list) + 2}"

        # Cap the total percentage at 100%
        row['Total Percentage'] = f"{min(100, (total_hours_worked / TOTAL_WORK_HOURS) * 100):.2f}%"
        template_data_list.append(row)

    # Convert the less than 8 hours data list to a DataFrame
    less_than_8_df = pd.DataFrame(less_than_8_data_list)


    # Identify discrepancies where Start Date and End Date do not match
    discrepancy_df = clockify_data[clockify_data['Start Date'] != clockify_data['End Date']]

    # Remove unnecessary columns from the discrepancy data
    discrepancy_df = discrepancy_df.drop(columns=['Client', 'Description', 'Task', 'Tags', 'Start Time', 'End Time', 'Duration (decimal)'])
    output_path = 'Clockify_Output_Final.xlsx'
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        output_template.to_excel(writer, index=False, sheet_name='Template')
        pivot_data.reset_index().to_excel(writer, index=False, sheet_name='Pivot Data')
        if not overtime_df.empty:
            overtime_df.to_excel(writer, index=False, sheet_name='Overtime')
        if not less_than_8_df.empty:
            less_than_8_df.to_excel(writer, index=False, sheet_name='Less than 8')
        if not discrepancy_df.empty:
            discrepancy_df.to_excel(writer, index=False, sheet_name='Discrepancy Data')

    # Beautify the Template, Overtime, Less than 8, and Discrepancy sheets using openpyxl
    wb = load_workbook(output_path)
    ws_template = wb['Template']

    # Set header formatting for the Template sheet
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for cell in ws_template[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        cell.border = thin_border

    # Set column widths and apply borders for the Template sheet
    for col in ws_template.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                cell_value = str(cell.value)  # Convert to string to avoid the TypeError
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except TypeError:
                continue
        adjusted_width = (max_length + 2)
        ws_template.column_dimensions[column].width = adjusted_width
        for cell in col:
            cell.border = thin_border

    # Apply similar formatting to the Overtime, Less than 8, and Discrepancy sheets if they exist
    for sheet_name in ['Overtime', 'Less than 8', 'Discrepancy Data']:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter # Get the column name
                for cell in col:
                    try:
                        cell_value = str(cell.value)  # Convert to string to avoid the TypeError
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                    except TypeError:
                        continue
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
                for cell in col:
                    cell.border = thin_border

    # Save the workbook with the updated formatting
    wb.save(output_path)
    return output_path

# Function to handle Block 3 logic
def block_3(clockify_data, output_path):
    # Load the existing Excel output file
    wb = load_workbook(output_path)

    # Extract the month name from the data
    start_date = clockify_data['Start Date'].min()
    month_name = calendar.month_name[start_date.month]

    # Add the original data as a new sheet
    ws_name = f"Clockify Data {month_name}"
    if ws_name in wb.sheetnames:
        ws = wb[ws_name]
    else:
        ws = wb.create_sheet(title=ws_name)

    # Write the column headers to the new sheet
    for col_num, column_title in enumerate(clockify_data.columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Write the DataFrame data to the new sheet
    for r_idx, row in enumerate(clockify_data.values, 2):  # Starting from row 2 to account for headers
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Apply highlighting to weekend records
    weekend_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for r_idx in range(2, ws.max_row + 1):  # Start from 2 because the first row has headers
        start_date = ws.cell(row=r_idx, column=clockify_data.columns.get_loc('Start Date') + 1).value
        if start_date.weekday() >= 5:  # Saturday or Sunday
            for c_idx in range(1, clockify_data.shape[1] + 1):
                ws.cell(row=r_idx, column=c_idx).fill = weekend_fill

    # Set column widths and apply borders
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column letter
        for cell in col:
            try:
                cell_value = str(cell.value)
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except TypeError:
                continue
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
        for cell in col:
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Save the workbook with the new sheet added
    wb.save(output_path)
    return output_path

# Streamlit app
st.title("Clockify Data Processing App")

uploaded_file = st.file_uploader("Upload your Clockify Excel data file", type="xlsx")

if uploaded_file:
    # Load the Clockify data
    clockify_data = pd.read_excel(uploaded_file)

    # Execute the blocks in sequence
    output_template, pivot_data, overtime_df, less_than_8_df, discrepancy_df, clockify_data = block_1(clockify_data)
    output_path = block_2(output_template, pivot_data, overtime_df, less_than_8_df, discrepancy_df)
    final_output_path = block_3(clockify_data, output_path)

    # Provide download link for the final Excel file
    with open(final_output_path, "rb") as file:
        st.download_button(
            label="Download processed Clockify data",
            data=file,
            file_name=final_output_path,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    st.success("Data processing complete. You can download the file now.")
else:
    st.info("Please upload the Clockify data file to proceed.")
